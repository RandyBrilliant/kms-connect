"""
Excel export filtered by recruitment activity dates (pra-seleksi, interview, medical, psikotes).

One row per matching JobApplication, prefixed with lamaran/tahapan columns and followed
by the full pelamar export columns from account.services.export.generate_applicants_excel.
"""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any, Iterable

from django.db.models import Q, QuerySet
from django.utils import timezone

from account.models import CustomUser, UserRole
from account.services.export import (
    EXPORT_SELECT_RELATED_APPLICANT_PROFILE_REGIONS,
    generate_applicants_excel,
)
from main.models import ApplicationStatus, JobApplication
from main.services import ApplicationService
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ACTIVITY_PRA_SELEKSI = "PRA_SELEKSI"
ACTIVITY_INTERVIEW = "INTERVIEW"
ACTIVITY_MEDICAL = "MEDICAL"
ACTIVITY_PSIKOTES = "PSIKOTES"

VALID_ACTIVITIES = {
    ACTIVITY_PRA_SELEKSI,
    ACTIVITY_INTERVIEW,
    ACTIVITY_MEDICAL,
    ACTIVITY_PSIKOTES,
}

ACTIVITY_LABELS = {
    ACTIVITY_PRA_SELEKSI: "Pra-Seleksi",
    ACTIVITY_INTERVIEW: "Interview",
    ACTIVITY_MEDICAL: "Medical",
    ACTIVITY_PSIKOTES: "Psikotes",
}

LAMARAN_EXPORT_COLUMNS = [
    ("Aktivitas (Filter)", "activity_filter"),
    ("Lowongan", "job_title"),
    ("Perusahaan", "company_name"),
    ("Status Lamaran", "status_label"),
    ("Grup Tahapan", "stage_group"),
    ("Nama Tahapan Pra-Seleksi", "batch_name"),
    ("Label Tahap Pra-Seleksi", "batch_tahap_label"),
    ("Sesi Interview", "cohort_name"),
    ("Sub-tahapan Diterima", "diterima_step_label"),
    ("Tgl. Pra-Seleksi", "pra_seleksi_date"),
    ("Tgl. Interview", "interview_date"),
    ("Tgl. Medical", "tgl_medical"),
    ("Tgl. FWCMS & Psikotes", "tgl_fwcm_psikotes"),
]

_DITERIMA_STEP_LABELS = dict(ApplicationService.DOCUMENT_COLLECTION_STEP_ORDER)


def _parse_date(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return date.fromisoformat(value.strip()[:10])
    except (TypeError, ValueError):
        return None


def _date_in_range(value: date | datetime | None, start: date, end: date) -> bool:
    if value is None:
        return False
    if isinstance(value, datetime):
        if timezone.is_aware(value):
            value = timezone.localtime(value).date()
        else:
            value = value.date()
    return start <= value <= end


def _stage_group_for_status(status: str) -> str:
    if status == ApplicationStatus.PRA_SELEKSI:
        return "Pra-Seleksi"
    if status in (ApplicationStatus.INTERVIEW, ApplicationStatus.CADANGAN):
        return "Interview"
    if status == ApplicationStatus.DITERIMA:
        return "Diterima"
    if status == ApplicationStatus.BERANGKAT:
        return "Berangkat"
    if status == ApplicationStatus.SELESAI:
        return "Selesai"
    if status == ApplicationStatus.DITOLAK:
        return "Ditolak"
    if status == ApplicationStatus.TRANSFERRED:
        return "Dipindah Lowongan"
    return status or "-"


def _fmt_date(value: date | datetime | None) -> str:
    if value is None:
        return "-"
    if isinstance(value, datetime):
        if timezone.is_aware(value):
            value = timezone.localtime(value)
        return value.strftime("%d-%m-%Y %H:%M")
    return value.strftime("%d-%m-%Y")


def matched_activities_for_application(
    app: JobApplication,
    activities: set[str],
    date_from: date,
    date_to: date,
) -> list[str]:
    matched: list[str] = []
    profile = app.applicant

    if ACTIVITY_PRA_SELEKSI in activities and app.batch_id:
        batch_date = getattr(app.batch, "pra_seleksi_date", None)
        if _date_in_range(batch_date, date_from, date_to):
            matched.append(ACTIVITY_LABELS[ACTIVITY_PRA_SELEKSI])

    if ACTIVITY_INTERVIEW in activities:
        cohort_date = (
            app.interview_cohort.interview_date
            if app.interview_cohort_id
            else None
        )
        batch_interview = (
            app.batch.interview_date if app.batch_id and not app.interview_cohort_id else None
        )
        interview_date = cohort_date or batch_interview
        if _date_in_range(interview_date, date_from, date_to):
            matched.append(ACTIVITY_LABELS[ACTIVITY_INTERVIEW])

    if ACTIVITY_MEDICAL in activities and profile:
        if _date_in_range(getattr(profile, "tgl_medical", None), date_from, date_to):
            matched.append(ACTIVITY_LABELS[ACTIVITY_MEDICAL])

    if ACTIVITY_PSIKOTES in activities and profile:
        if _date_in_range(
            getattr(profile, "tgl_fwcm_psikotes", None), date_from, date_to
        ):
            matched.append(ACTIVITY_LABELS[ACTIVITY_PSIKOTES])

    return matched


def build_activity_application_queryset(
    activities: Iterable[str],
    date_from: date,
    date_to: date,
) -> QuerySet[JobApplication]:
    activity_set = {a for a in activities if a in VALID_ACTIVITIES}
    if not activity_set:
        return JobApplication.objects.none()

    q = Q()
    if ACTIVITY_PRA_SELEKSI in activity_set:
        q |= Q(
            batch__pra_seleksi_date__date__gte=date_from,
            batch__pra_seleksi_date__date__lte=date_to,
        )
    if ACTIVITY_INTERVIEW in activity_set:
        q |= Q(
            interview_cohort__interview_date__date__gte=date_from,
            interview_cohort__interview_date__date__lte=date_to,
        )
        q |= Q(
            interview_cohort__isnull=True,
            batch__interview_date__date__gte=date_from,
            batch__interview_date__date__lte=date_to,
        )
    if ACTIVITY_MEDICAL in activity_set:
        q |= Q(
            applicant__tgl_medical__gte=date_from,
            applicant__tgl_medical__lte=date_to,
        )
    if ACTIVITY_PSIKOTES in activity_set:
        q |= Q(
            applicant__tgl_fwcm_psikotes__gte=date_from,
            applicant__tgl_fwcm_psikotes__lte=date_to,
        )

    return (
        JobApplication.objects.filter(q)
        .select_related(
            "applicant",
            "applicant__user",
            "job",
            "job__company",
            "batch",
            "interview_cohort",
        )
        .distinct()
        .order_by("applicant__user__full_name", "job__title", "id")
    )


def _lamaran_row_values(
    app: JobApplication,
    activities: set[str],
    date_from: date,
    date_to: date,
) -> dict[str, str]:
    profile = app.applicant
    batch = app.batch
    cohort = app.interview_cohort
    diterima_code = getattr(app, "diterima_current_step", None) or ""
    diterima_label = _DITERIMA_STEP_LABELS.get(diterima_code, diterima_code or "-")

    interview_date = None
    if cohort and cohort.interview_date:
        interview_date = cohort.interview_date
    elif batch and batch.interview_date:
        interview_date = batch.interview_date

    return {
        "activity_filter": ", ".join(
            matched_activities_for_application(app, activities, date_from, date_to)
        ),
        "job_title": app.job.title if app.job_id else "-",
        "company_name": (
            app.job.company.company_name
            if app.job_id and getattr(app.job, "company", None)
            else "-"
        ),
        "status_label": app.get_status_display() if app.status else "-",
        "stage_group": _stage_group_for_status(app.status),
        "batch_name": batch.name if batch else "-",
        "batch_tahap_label": batch.display_tahap_label if batch else "-",
        "cohort_name": cohort.name if cohort else "-",
        "diterima_step_label": diterima_label if app.status == ApplicationStatus.DITERIMA else "-",
        "pra_seleksi_date": _fmt_date(batch.pra_seleksi_date if batch else None),
        "interview_date": _fmt_date(interview_date),
        "tgl_medical": _fmt_date(getattr(profile, "tgl_medical", None)),
        "tgl_fwcm_psikotes": _fmt_date(getattr(profile, "tgl_fwcm_psikotes", None)),
    }


def generate_activity_applications_excel(
    applications: Iterable[JobApplication],
    activities: Iterable[str],
    date_from: date,
    date_to: date,
    request: Any = None,
) -> BytesIO:
    apps = list(applications)
    activity_set = {a for a in activities if a in VALID_ACTIVITIES}

    user_ids = list({app.applicant.user_id for app in apps if app.applicant_id})
    applicants_qs = (
        CustomUser.objects.filter(id__in=user_ids, role=UserRole.APPLICANT)
        .select_related(*EXPORT_SELECT_RELATED_APPLICANT_PROFILE_REGIONS)
        .prefetch_related(
            "applicant_profile__work_experiences",
            "applicant_profile__documents__document_type",
            "applicant_profile__documents__reviewed_by",
        )
        .order_by("full_name")
    )

    pelamar_excel = generate_applicants_excel(applicants_qs, request)
    src_wb = load_workbook(pelamar_excel)
    src_ws = src_wb.active

    src_headers = [cell.value for cell in src_ws[1]]
    nik_col_idx = None
    for idx, header in enumerate(src_headers):
        if header == "NIK":
            nik_col_idx = idx
            break

    rows_by_nik: dict[str, tuple] = {}
    rows_by_user_id: dict[int, tuple] = {}
    user_list = list(applicants_qs)
    for row_idx, row in enumerate(
        src_ws.iter_rows(min_row=2, max_row=src_ws.max_row, values_only=True), start=2
    ):
        if not row or all(v is None for v in row):
            continue
        if nik_col_idx is not None:
            nik_val = row[nik_col_idx]
            if nik_val:
                rows_by_nik[str(nik_val).strip()] = row
        user_pos = row_idx - 2
        if user_pos < len(user_list):
            rows_by_user_id[user_list[user_pos].id] = row

    dst_wb = Workbook()
    dst_ws = dst_wb.active
    dst_ws.title = "Export Aktivitas"

    all_headers = [label for label, _ in LAMARAN_EXPORT_COLUMNS] + list(src_headers)

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_alignment = Alignment(vertical="top", wrap_text=True)

    for col_idx, label in enumerate(all_headers, start=1):
        cell = dst_ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    row_num = 2
    for app in apps:
        lamaran = _lamaran_row_values(app, activity_set, date_from, date_to)
        if not lamaran["activity_filter"]:
            continue

        profile = app.applicant
        pelamar_row = None
        if profile and profile.nik:
            pelamar_row = rows_by_nik.get(str(profile.nik).strip())
        if pelamar_row is None and app.applicant.user_id:
            pelamar_row = rows_by_user_id.get(app.applicant.user_id)
        if pelamar_row is None:
            pelamar_row = tuple("-" for _ in src_headers)

        for col_idx, (_, field_key) in enumerate(LAMARAN_EXPORT_COLUMNS, start=1):
            cell = dst_ws.cell(row=row_num, column=col_idx, value=lamaran.get(field_key, "-"))
            cell.alignment = data_alignment

        offset = len(LAMARAN_EXPORT_COLUMNS)
        for col_idx, value in enumerate(pelamar_row, start=1):
            cell = dst_ws.cell(row=row_num, column=offset + col_idx, value=value)
            cell.alignment = data_alignment

        row_num += 1

    for col_idx, label in enumerate(all_headers, start=1):
        col_letter = get_column_letter(col_idx)
        min_width = max(len(str(label)) + 2, 10)
        if col_idx <= len(LAMARAN_EXPORT_COLUMNS):
            min_width = max(min_width, 14)
        dst_ws.column_dimensions[col_letter].width = min_width

    dst_ws.freeze_panes = "A2"

    output = BytesIO()
    dst_wb.save(output)
    output.seek(0)
    return output


def parse_activity_export_params(
    query_params,
) -> tuple[set[str], date, date]:
    """Parse and validate query params for activity export."""
    activities = query_params.getlist("activity")
    if not activities:
        single = query_params.get("activity")
        if single:
            activities = [single]
    activity_set = {a.strip().upper() for a in activities if a.strip()}
    unknown = activity_set - VALID_ACTIVITIES
    if unknown:
        raise ValueError(
            f"Aktivitas tidak valid: {', '.join(sorted(unknown))}. "
            f"Pilihan: {', '.join(sorted(VALID_ACTIVITIES))}."
        )
    if not activity_set:
        raise ValueError("Pilih minimal satu aktivitas.")

    date_from = _parse_date(query_params.get("date_from"))
    date_to = _parse_date(query_params.get("date_to"))
    if not date_from or not date_to:
        raise ValueError("Parameter date_from dan date_to wajib diisi (YYYY-MM-DD).")
    if date_from > date_to:
        raise ValueError("date_from tidak boleh setelah date_to.")

    return activity_set, date_from, date_to
