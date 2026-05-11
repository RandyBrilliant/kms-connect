"""
Excel export utilities for ApplicantProfile (pelamar).

Design goals:
- Pure functions for Excel generation (no DB queries here).
- Reusable export logic that can be called from views or tasks.
- Handles large datasets efficiently using streaming.
- Follows the same patterns as scoring.py (services separation).
- Includes all data: basic info, work experiences, documents with downloadable links.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any, Iterable
from datetime import date, datetime

from django.conf import settings

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Excel column headers follow client-provided template order.
# NextOfKinRelationship imported lazily in spouse/wali helpers to avoid import cycles.
EXPORT_COLUMNS = [
    ("TANGGAL DAFTAR", "registration_date"),
    ("Pemberi Rujukan", "referrer_name"),
    ("NIK", "nik"),
    ("Nama", "full_name"),
    ("Tempat Lahir", "birth_place"),
    ("Tanggal Lahir", "birth_date"),
    # Alamat pelamar: baris gabungan jalan + kelurahan + kecamatan; kab/kota & provinsi terpisah
    ("Alamat", "address_combined"),
    ("Kabupaten/Kota", "district_display"),
    ("Provinsi", "province_display"),
    ("Kode Pos", "postal_code"),
    ("No. HP", "contact_phone"),
    ("Email", "email"),
    ("JUMLAH SAUDARA", "sibling_count"),
    ("ANDA ANAK KEBERAPA", "birth_order"),
    ("PENGALAMAN KERJA 1 (TULISKAN NAMA PERUSAHAAN DAN KOTANYA) ", "work_company_1"),
    ("NEGARA TEMPAT ANDA BEKERJA TERSEBUT DIMANA (1) ?", "work_country_1"),
    ("JABATAN SEBAGAI 1 (CONTOH : OPERATOR / LEADER /  SUPERVISOR, PENJAGA TOKO, DLL)", "work_position_1"),
    ("MASA BEKERJA 1 (Contoh Januari 2022 - Januari 2024)", "work_period_1"),
    ("PENGALAMAN KERJA 2 (TULISKAN NAMA PERUSAHAAN DAN KOTANYA) ", "work_company_2"),
    ("NEGARA TEMPAT ANDA BEKERJA TERSEBUT DIMANA (2) ?", "work_country_2"),
    ("JABATAN SEBAGAI 2 (CONTOH : OPERATOR / LEADER /  SUPERVISOR, PENJAGA TOKO, DLL)", "work_position_2"),
    ("MASA BEKERJA 2 (Contoh Januari 2022 - Januari 2024)", "work_period_2"),
    ("NAMA BAPAK", "father_name"),
    ("PEKERJAAN ", "father_occupation"),
    ("UMUR", "father_age"),
    ("NAMA IBU", "mother_name"),
    ("PEKERJAAN ", "mother_occupation"),
    ("UMUR", "mother_age"),
    ("NAMA SUAMI", "spouse_name"),
    ("PEKERJAAN ", "spouse_occupation"),
    ("UMUR", "spouse_age"),
    # Alamat orang tua / keluarga (satu blok wilayah di profil)
    ("Alamat Keluarga (Orang Tua)", "family_address_combined"),
    ("Kabupaten/Kota Keluarga", "family_district_display"),
    ("Provinsi Keluarga", "family_province_display"),
    ("Kode Pos Keluarga", "family_postal_code"),
    # Suami/istri atau wali: kolom sama dipakai jika ada nama pasangan / ahli waris terkait & ada data alamat keluarga
    ("Alamat Suami/Istri atau Wali", "spouse_wali_address_combined"),
    ("Kabupaten/Kota (Suami/Istri/Wali)", "spouse_wali_district_display"),
    ("Provinsi (Suami/Istri/Wali)", "spouse_wali_province_display"),
    ("Kode Pos (Suami/Istri/Wali)", "spouse_wali_postal_code"),
    ("No. HP Keluarga", "family_phone"),
    ("Email Keluarga", "family_email"),
    ("AGAMA", "religion"),
    ("Pendidikan", "education_level"),
    ("Jurusan", "education_major"),
    ("TINGGI", "height_cm"),
    ("BERAT", "weight_kg"),
    ("APAKAH ANDA MEMAKAI KACAMATA (MATA MINUS)", "wears_glasses"),
    ("ANDA MENULIS DENGAN TANGAN ?", "writing_hand"),
    ("STATUS PERKAWINAN ANDA", "marital_status"),
    ("APAKAH ANDA SUDAH MEMILIKI PASPOR ?", "has_passport"),
    ("ISI NOMOR PASPOR ANDA", "passport_number"),
    ("TANGGAL BERAKHIR MASA BERLAKU PASPOR", "passport_expiry_date"),
    ("Jenis Kelamin", "gender"),
    # Admin
    ("Status Verifikasi", "verification_status"),
    ("Diverifikasi Oleh", "verified_by_name"),
    ("Tanggal Verifikasi", "verified_at"),
    ("Catatan Verifikasi", "verification_notes"),
    ("Skor", "score"),
    # Admin process & finance (ApplicantProfile admin-only fields)
    ("Tgl. Medical", "tgl_medical"),
    ("Hasil Medical", "hasil_medical"),
    ("Tgl. Bayar SML", "tgl_bayar_sml"),
    ("Tgl. FWCMS & Psikotes", "tgl_fwcm_psikotes"),
    ("Tgl. Bayar Psikotes", "tgl_bayar_psikotes"),
    ("Tgl. Bayar BPJS Pra", "tgl_bayar_bpjs_pra"),
    ("Tgl. Bayar BPJS Purna", "tgl_bayar_bpjs_purna"),
    ("No. ID SISKO", "no_id_sisko"),
    ("Disnaker", "disnaker"),
    ("No. SIP", "no_sip"),
    ("No. JO", "no_jo"),
    ("Biaya Ready Paspor", "biaya_ready_paspor"),
    ("Pengembalian Biaya", "pengembalian_biaya"),
    ("Tgl. Pengembalian", "tgl_pengembalian"),
    ("Jlh Uang Transport", "jlh_uang_transport"),
    ("Bank", "bank"),
    ("No. Rekening", "no_rek"),
    ("Tanggal Pengembalian (Transfer)", "tanggal_pengembalian"),
    ("Tgl. Kirim Bio ke MY", "tgl_kirim_bio_ke_mly"),
    ("Tgl. Calling Visa", "tgl_calling_visa"),
    ("No. Calling Visa", "no_calling_visa"),
    # Account
    ("Status Akun", "is_active"),
    ("Email Terverifikasi", "email_verified"),
    ("Tanggal Bergabung", "created_at"),
    # Work Experiences (concatenated)
    ("Pengalaman Kerja", "work_experiences"),
    # Documents will be added dynamically per document type
]

# Used by list/export querysets to avoid N+1 when resolving wilayah from Village FKs.
EXPORT_SELECT_RELATED_APPLICANT_PROFILE_REGIONS = (
    "applicant_profile__province",
    "applicant_profile__district",
    "applicant_profile__village__district__regency__province",
    "applicant_profile__family_province",
    "applicant_profile__family_district",
    "applicant_profile__family_village__district__regency__province",
    "applicant_profile__referrer",
    "applicant_profile__verified_by",
)


def _get_nested_value(obj: Any, path: str, default: str = "-") -> str:
    """
    Safely get a nested attribute value and convert to string.
    
    Supports:
    - Dotted paths (e.g., "user.full_name")
    - Direct attributes
    - Returns default if value is None or empty
    """
    if not path:
        return default
    
    current: Any = obj
    for part in path.split("."):
        if current is None:
            return default
        if isinstance(current, dict):
            current = current.get(part)
        else:
            current = getattr(current, part, None)
    
    if current is None:
        return default
    if isinstance(current, bool):
        return "Ya" if current else "Tidak"
    if isinstance(current, datetime):
        return current.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(current, (int, float)):
        return str(current)
    
    return str(current) if current else default


def _format_verification_status(status: str | None) -> str:
    """Format verification status for display."""
    status_map = {
        "DRAFT": "Draft",
        "SUBMITTED": "Dikirim",
        "ACCEPTED": "Diterima",
        "REJECTED": "Ditolak",
    }
    return status_map.get(status or "", status or "-")


def _format_gender(gender: str | None) -> str:
    """Format gender for display."""
    gender_map = {
        "M": "Laki-laki",
        "F": "Perempuan",
    }
    return gender_map.get(gender or "", gender or "-")


def _format_education_level(level: str | None) -> str:
    """Format education level for display."""
    level_map = {
        "SMP": "SMP",
        "SMA": "SMA",
        "SMK": "SMK",
        "MA": "MA",
        "D3": "D3",
        "S1": "S1",
    }
    return level_map.get(level or "", level or "-")


def _format_industry_type(industry: str | None) -> str:
    """Format industry type for display."""
    industry_map = {
        "SEMICONDUCTOR": "Semiconductor",
        "ELECTRONICS": "Elektronik",
        "OTHER_FACTORY": "Pabrik Lain",
        "SERVICES": "Jasa",
        "OTHER": "Lain Lain",
        "NEVER_WORKED": "Belum Pernah Bekerja",
    }
    return industry_map.get(industry or "", industry or "-")


def _format_document_review_status(status: str | None) -> str:
    """Format document review status for display."""
    status_map = {
        "PENDING": "Pending",
        "APPROVED": "Disetujui",
        "REJECTED": "Ditolak",
    }
    return status_map.get(status or "", status or "-")


def _format_date_dmy(value: Any) -> str:
    """Format date/datetime as DD-MM-YYYY for export."""
    if not value:
        return "-"
    if isinstance(value, datetime):
        return value.strftime("%d-%m-%Y")
    if isinstance(value, date):
        return value.strftime("%d-%m-%Y")
    return str(value)


def _format_date_dd_mon_yyyy(value: Any) -> str:
    """Format date/datetime as DD-Mon-YYYY (e.g. 14-Mar-2012)."""
    if not value:
        return "-"

    dt_value: date | None = None
    if isinstance(value, datetime):
        dt_value = value.date()
    elif isinstance(value, date):
        dt_value = value
    elif isinstance(value, str):
        try:
            dt_value = datetime.fromisoformat(value).date()
        except ValueError:
            return value
    else:
        return str(value)

    month_abbr = [
        "Jan",
        "Feb",
        "Mar",
        "Apr",
        "May",
        "Jun",
        "Jul",
        "Aug",
        "Sep",
        "Oct",
        "Nov",
        "Dec",
    ]
    return f"{dt_value.day:02d}-{month_abbr[dt_value.month - 1]}-{dt_value.year}"


def _safe_name(obj: Any) -> str:
    if not obj:
        return "-"
    return getattr(obj, "name", None) or str(obj) or "-"


def _country_full_name(value: Any) -> str:
    """
    Convert CountryField value into full country name.
    Falls back safely to string representation.
    """
    if not value:
        return "-"
    name = getattr(value, "name", None)
    if name:
        return str(name)
    code = getattr(value, "code", None)
    if code and hasattr(value, "name"):
        return str(value.name)
    return str(value)


def _first_non_empty(*values: Any) -> str:
    for value in values:
        text = str(value).strip() if value is not None else ""
        if text:
            return text
    return "-"


def _kecamatan_name_from_village(village: Any) -> str:
    if not village:
        return ""
    dist = getattr(village, "district", None)
    return str(getattr(dist, "name", "") or "").strip()


def _kelurahan_name_from_village(village: Any) -> str:
    return str(getattr(village, "name", "") or "").strip() if village else ""


def _combined_street_kel_kec(street: str, village: Any) -> str:
    """Alamat baris: teks jalan, kelurahan, kecamatan (dipisah koma)."""
    parts: list[str] = []
    s = (street or "").strip()
    if s:
        parts.append(s)
    kel = _kelurahan_name_from_village(village)
    if kel:
        parts.append(kel)
    kec = _kecamatan_name_from_village(village)
    if kec:
        parts.append(kec)
    return ", ".join(parts) if parts else "-"


def _has_family_address_data(profile: Any) -> bool:
    if (getattr(profile, "family_address", None) or "").strip():
        return True
    if getattr(profile, "family_village_id", None):
        return True
    if getattr(profile, "family_district_id", None):
        return True
    if getattr(profile, "family_province_id", None):
        return True
    return False


def _show_spouse_wali_address_block(profile: Any) -> bool:
    """
    Tampilkan blok suami/istri/wali bila ada data alamat keluarga dan konteks pasangan/wali.
    Orang tua (Ayah/Ibu) hanya di blok 'Alamat Keluarga'; tidak diulang di sini.
    """
    if not profile or not _has_family_address_data(profile):
        return False
    if (getattr(profile, "spouse_name", None) or "").strip():
        return True
    heir = (getattr(profile, "heir_name", None) or "").strip()
    rel = getattr(profile, "heir_relationship", None) or ""
    if not heir:
        return False
    if rel in ("AYAH", "IBU"):
        return False
    return rel in (
        "SUAMI",
        "ISTRI",
        "KAKAK",
        "ADIK",
        "ANAK",
        "PAMAN",
        "BIBI",
        "LAINNYA",
    )


def _work_experience_at(profile: Any, index: int):
    work_experiences = getattr(profile, "work_experiences", None)
    if not work_experiences:
        return None
    experiences = list(work_experiences.all()) if hasattr(work_experiences, "all") else list(work_experiences)
    if not experiences or index < 0 or index >= len(experiences):
        return None
    return experiences[index]


def _get_region_display(profile: Any, field: str) -> str:
    """
    Extract region display string from village_display or family_village_display.
    Also handles building from related objects if display is not available.
    """
    display_obj = getattr(profile, field, None)
    
    # Handle dict-like objects (from serializer)
    if isinstance(display_obj, dict):
        parts = []
        if display_obj.get("province"):
            parts.append(display_obj["province"])
        if display_obj.get("regency"):
            parts.append(display_obj["regency"])
        if display_obj.get("district"):
            parts.append(display_obj["district"])
        if display_obj.get("village"):
            parts.append(display_obj["village"])
        return ", ".join(parts) if parts else "-"
    
    # Handle object with attributes
    if display_obj:
        parts = []
        if hasattr(display_obj, "province") and display_obj.province:
            parts.append(str(display_obj.province))
        if hasattr(display_obj, "regency") and display_obj.regency:
            parts.append(str(display_obj.regency))
        if hasattr(display_obj, "district") and display_obj.district:
            parts.append(str(display_obj.district))
        if hasattr(display_obj, "village") and display_obj.village:
            parts.append(str(display_obj.village))
        if parts:
            return ", ".join(parts)
    
    # Fallback: build from related objects (if select_related was used)
    if field == "village_display":
        parts = []
        province = getattr(profile, "province", None)
        district = getattr(profile, "district", None)
        village = getattr(profile, "village", None)
        
        if province:
            parts.append(getattr(province, "name", str(province)))
        if district:
            parts.append(getattr(district, "name", str(district)))
        if village:
            parts.append(getattr(village, "name", str(village)))
        
        return ", ".join(parts) if parts else "-"
    
    return "-"


def _get_file_url(file_field: Any, request: Any = None) -> str:
    """
    Get full URL for a file field.
    
    Args:
        file_field: Django FileField
        request: Django request object (optional, for building absolute URLs)
    
    Returns:
        Full URL to the file or "-" if no file
    """
    if not file_field:
        return "-"
    
    try:
        # Get the file URL
        url = file_field.url if hasattr(file_field, 'url') else str(file_field)
        
        # If URL is relative, make it absolute
        if url.startswith('/'):
            # Use MEDIA_URL from settings
            media_url = getattr(settings, 'MEDIA_URL', '/media/')
            if media_url.startswith('http'):
                # Already absolute (e.g., S3/CDN)
                return url
            else:
                # Build absolute URL
                # Try to get domain from request
                if request:
                    scheme = 'https' if request.is_secure() else 'http'
                    host = request.get_host()
                    return f"{scheme}://{host}{url}"
                else:
                    # Fallback: just return the relative URL
                    return url
        
        return url
    except Exception:
        return "-"


def _format_work_experiences(profile: Any, request: Any = None) -> str:
    """
    Format all work experiences into a readable string.
    
    Returns:
        Multi-line string with all work experiences or "-" if none
    """
    work_experiences = getattr(profile, "work_experiences", None)
    if not work_experiences:
        return "-"
    
    try:
        # Get all work experiences (assuming it's a related manager or list)
        experiences = list(work_experiences.all()) if hasattr(work_experiences, 'all') else work_experiences
        
        if not experiences:
            return "-"
        
        result = []
        for idx, exp in enumerate(experiences, 1):
            parts = [f"#{idx}"]
            
            company = getattr(exp, "company_name", None)
            if company:
                parts.append(f"Perusahaan: {company}")
            
            position = getattr(exp, "position", None)
            if position:
                parts.append(f"Jabatan: {position}")
            
            department = getattr(exp, "department", None)
            if department:
                parts.append(f"Bagian: {department}")
            
            location = getattr(exp, "location", None)
            if location:
                parts.append(f"Lokasi: {location}")
            
            country = getattr(exp, "country", None)
            if country:
                parts.append(f"Negara: {_country_full_name(country)}")
            
            industry = getattr(exp, "industry_type", None)
            if industry:
                parts.append(f"Industri: {_format_industry_type(industry)}")
            
            start_date = getattr(exp, "start_date", None)
            end_date = getattr(exp, "end_date", None)
            if start_date:
                date_str = f"Periode: {start_date.strftime('%Y-%m-%d')}"
                if end_date:
                    date_str += f" s/d {end_date.strftime('%Y-%m-%d')}"
                else:
                    date_str += " s/d sekarang"
                parts.append(date_str)
            
            result.append(" | ".join(parts))
        
        return "\n".join(result)
    except Exception as e:
        return f"Error: {str(e)}"


def _format_documents(profile: Any, request: Any = None) -> str:
    """
    Format all documents into a readable string with downloadable links.
    
    Returns:
        Multi-line string with all documents and links or "-" if none
    """
    documents = getattr(profile, "documents", None)
    if not documents:
        return "-"
    
    try:
        # Get all documents (assuming it's a related manager or list)
        docs = list(documents.all()) if hasattr(documents, 'all') else documents
        
        if not docs:
            return "-"
        
        result = []
        for doc in docs:
            parts = []
            
            # Document type
            doc_type = getattr(doc, "document_type", None)
            if doc_type:
                type_name = getattr(doc_type, "name", str(doc_type))
                parts.append(f"Tipe: {type_name}")
            
            # Review status
            review_status = getattr(doc, "review_status", None)
            if review_status:
                parts.append(f"Status: {_format_document_review_status(review_status)}")
            
            # File URL
            file_field = getattr(doc, "file", None)
            if file_field:
                url = _get_file_url(file_field, request)
                parts.append(f"Link: {url}")
            
            # Reviewed by
            reviewed_by = getattr(doc, "reviewed_by", None)
            if reviewed_by:
                reviewer_name = getattr(reviewed_by, "full_name", None) or getattr(reviewed_by, "email", "")
                if reviewer_name:
                    parts.append(f"Direview: {reviewer_name}")
            
            # Upload date
            uploaded_at = getattr(doc, "uploaded_at", None)
            if uploaded_at:
                parts.append(f"Tanggal: {uploaded_at.strftime('%Y-%m-%d %H:%M')}")
            
            if parts:
                result.append(" | ".join(parts))
        
        return "\n".join(result)
    except Exception as e:
        return f"Error: {str(e)}"


def generate_applicants_excel(applicants: Iterable[Any], request: Any = None) -> BytesIO:
    """
    Generate Excel file from applicants queryset/iterable.
    
    Args:
        applicants: Iterable of CustomUser objects with applicant_profile loaded.
                   Should use select_related("applicant_profile__user") and
                   prefetch_related("applicant_profile__work_experiences", "applicant_profile__documents")
                   for optimal performance.
        request: Django request object for building absolute URLs
    
    Returns:
        BytesIO object containing the Excel file.
    
    Design:
    - Pure function (no DB queries, no side effects).
    - Efficient for large datasets (streams data row by row).
    - Uses openpyxl for Excel generation.
    - Includes all data: basic info, work experiences, documents with downloadable links.
    - Each document type gets its own column with clickable links.
    """
    from account.models import DocumentType
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Daftar Pelamar"
    
    # Get all document types and force the client-requested export order first.
    desired_document_order = [
        "KTP",
        "Ijazah",
        "Kartu Keluarga",
        "Paspor",
        "Pas Photo",
        "CV",
        "Surat Izin Keluarga (Form Biru)",
        "Sertifikat Keterampilan",
        "KTP Orangtua / Wali",
        "Surat Kesehatan",
        "Surat Keterangan Pemberi Izin",
        "Buku Nikah",
        "Perjanjian Penempatan",
        "Surat Keterangan Status Perkawinan",
        "Kartu BPJS Kesehatan",
    ]
    all_doc_types = list(DocumentType.objects.all().order_by("sort_order", "code"))
    doc_by_name = {d.name: d for d in all_doc_types}
    ordered = [doc_by_name[name] for name in desired_document_order if name in doc_by_name]
    remainder = [d for d in all_doc_types if d.name not in desired_document_order]
    document_types = ordered + remainder
    
    # Build full column list: base columns + document type columns
    all_columns = list(EXPORT_COLUMNS)
    document_type_columns = {}  # Map document_type.id to column index
    
    for doc_type in document_types:
        doc_type_label = doc_type.name
        doc_type_field = f"document_{doc_type.code}"
        all_columns.append((doc_type_label, doc_type_field))
        document_type_columns[doc_type.id] = len(all_columns) - 1
    
    # Header row styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Write headers
    for col_idx, (label, _) in enumerate(all_columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data row styling
    data_alignment = Alignment(vertical="top", wrap_text=True)
    
    # Write data rows
    row_num = 2
    for applicant in applicants:
        profile = getattr(applicant, "applicant_profile", None)
        if not profile:
            continue
        
        user = getattr(applicant, "user", applicant) if hasattr(applicant, "user") else applicant
        
        # Initialize row data dict for document columns
        row_data = {}
        
        # Process base columns
        for col_idx, (_, field_path) in enumerate(EXPORT_COLUMNS, start=1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.alignment = data_alignment
            
            # Handle special fields
            if field_path == "full_name":
                # full_name is on CustomUser, not ApplicantProfile
                value = _get_nested_value(applicant, "full_name")
            elif field_path == "email":
                value = _get_nested_value(applicant, "email")
            elif field_path == "registration_date":
                registration_date = getattr(profile, "registration_date", None)
                if not registration_date:
                    registration_date = getattr(profile, "created_at", None) or getattr(applicant, "date_joined", None)
                value = _format_date_dmy(registration_date)
            elif field_path == "nik":
                value = _get_nested_value(profile, "nik")
            elif field_path == "contact_phone":
                value = _get_nested_value(profile, "contact_phone")
            elif field_path == "birth_date":
                value = _format_date_dd_mon_yyyy(getattr(profile, "birth_date", None))
            elif field_path == "birth_place":
                t = (getattr(profile, "birth_place_text", None) or "").strip()
                if t:
                    value = t
                else:
                    value = _safe_name(getattr(profile, "birth_place", None))
            elif field_path == "gender":
                gender = _get_nested_value(profile, "gender", "")
                value = _format_gender(gender) if gender != "-" else "-"
            elif field_path == "address_combined":
                village = getattr(profile, "village", None)
                value = _combined_street_kel_kec(
                    getattr(profile, "address", None) or "",
                    village,
                )
            elif field_path == "province_display":
                # Province name only; fallback from village → regency → province
                province = getattr(profile, "province", None)
                if province:
                    value = getattr(province, "name", "-")
                else:
                    village = getattr(profile, "village", None)
                    reg = (
                        getattr(village.district, "regency", None)
                        if village and getattr(village, "district", None)
                        else None
                    )
                    prov = getattr(reg, "province", None) if reg else None
                    value = getattr(prov, "name", "-") if prov else "-"
            elif field_path == "district_display":
                # Kabupaten/Kota (Regency)
                district = getattr(profile, "district", None)
                if district:
                    value = _safe_name(district)
                else:
                    village = getattr(profile, "village", None)
                    reg = (
                        getattr(village.district, "regency", None)
                        if village and getattr(village, "district", None)
                        else None
                    )
                    value = _safe_name(reg)
            elif field_path == "postal_code":
                village = getattr(profile, "village", None)
                value = _first_non_empty(
                    getattr(profile, "postal_code", None),
                    str(getattr(village, "code", "") or "").strip(),
                )
            elif field_path == "education_level":
                level = _get_nested_value(profile, "education_level", "")
                value = _format_education_level(level) if level != "-" else "-"
            elif field_path == "education_institution":
                value = _get_nested_value(profile, "education_institution")
            elif field_path == "education_major":
                value = _get_nested_value(profile, "education_major")
            elif field_path == "sibling_count":
                value = _get_nested_value(profile, "sibling_count")
            elif field_path == "birth_order":
                value = _get_nested_value(profile, "birth_order")
            elif field_path == "work_company_1":
                exp = _work_experience_at(profile, 0)
                if exp:
                    company = getattr(exp, "company_name", "") or ""
                    city = getattr(exp, "location", "") or ""
                    value = f"{company} - {city}".strip(" -") or "-"
                else:
                    value = "-"
            elif field_path == "work_country_1":
                exp = _work_experience_at(profile, 0)
                value = _country_full_name(getattr(exp, "country", None)) if exp else "-"
            elif field_path == "work_position_1":
                exp = _work_experience_at(profile, 0)
                value = str(getattr(exp, "position", "") or "-") if exp else "-"
            elif field_path == "work_period_1":
                exp = _work_experience_at(profile, 0)
                if exp:
                    start = _format_date_dmy(getattr(exp, "start_date", None))
                    end_raw = getattr(exp, "end_date", None)
                    end = _format_date_dmy(end_raw) if end_raw else "Sekarang"
                    value = f"{start} - {end}" if start != "-" else "-"
                else:
                    value = "-"
            elif field_path == "work_company_2":
                exp = _work_experience_at(profile, 1)
                if exp:
                    company = getattr(exp, "company_name", "") or ""
                    city = getattr(exp, "location", "") or ""
                    value = f"{company} - {city}".strip(" -") or "-"
                else:
                    value = "-"
            elif field_path == "work_country_2":
                exp = _work_experience_at(profile, 1)
                value = _country_full_name(getattr(exp, "country", None)) if exp else "-"
            elif field_path == "work_position_2":
                exp = _work_experience_at(profile, 1)
                value = str(getattr(exp, "position", "") or "-") if exp else "-"
            elif field_path == "work_period_2":
                exp = _work_experience_at(profile, 1)
                if exp:
                    start = _format_date_dmy(getattr(exp, "start_date", None))
                    end_raw = getattr(exp, "end_date", None)
                    end = _format_date_dmy(end_raw) if end_raw else "Sekarang"
                    value = f"{start} - {end}" if start != "-" else "-"
                else:
                    value = "-"
            elif field_path == "father_name":
                value = _get_nested_value(profile, "father_name")
            elif field_path == "father_occupation":
                value = _get_nested_value(profile, "father_occupation")
            elif field_path == "father_age":
                value = _get_nested_value(profile, "father_age")
            elif field_path == "mother_name":
                value = _get_nested_value(profile, "mother_name")
            elif field_path == "mother_occupation":
                value = _get_nested_value(profile, "mother_occupation")
            elif field_path == "mother_age":
                value = _get_nested_value(profile, "mother_age")
            elif field_path == "spouse_name":
                value = _get_nested_value(profile, "spouse_name")
            elif field_path == "spouse_occupation":
                value = _get_nested_value(profile, "spouse_occupation")
            elif field_path == "spouse_age":
                value = _get_nested_value(profile, "spouse_age")
            elif field_path == "education_graduation_year":
                value = _get_nested_value(profile, "education_graduation_year")
            # Family Info (orang tua / keluarga)
            elif field_path == "family_address_combined":
                fv = getattr(profile, "family_village", None)
                value = _combined_street_kel_kec(
                    getattr(profile, "family_address", None) or "",
                    fv,
                )
            elif field_path == "family_province_display":
                family_province = getattr(profile, "family_province", None)
                if family_province:
                    value = getattr(family_province, "name", "-")
                else:
                    fv = getattr(profile, "family_village", None)
                    reg = (
                        getattr(fv.district, "regency", None)
                        if fv and getattr(fv, "district", None)
                        else None
                    )
                    prov = getattr(reg, "province", None) if reg else None
                    value = getattr(prov, "name", "-") if prov else "-"
            elif field_path == "family_district_display":
                fd = getattr(profile, "family_district", None)
                if fd:
                    value = _safe_name(fd)
                else:
                    fv = getattr(profile, "family_village", None)
                    reg = (
                        getattr(fv.district, "regency", None)
                        if fv and getattr(fv, "district", None)
                        else None
                    )
                    value = _safe_name(reg)
            elif field_path == "family_postal_code":
                family_village = getattr(profile, "family_village", None)
                value = _first_non_empty(
                    getattr(profile, "family_postal_code", None),
                    str(getattr(family_village, "code", "") or "").strip(),
                )
            elif field_path == "family_phone":
                value = _first_non_empty(
                    getattr(profile, "heir_contact_phone", None),
                    getattr(profile, "father_phone", None),
                    getattr(profile, "mother_phone", None),
                )
            elif field_path == "family_email":
                value = _get_nested_value(profile, "family_email")
            elif field_path == "spouse_wali_address_combined":
                if _show_spouse_wali_address_block(profile):
                    fv = getattr(profile, "family_village", None)
                    value = _combined_street_kel_kec(
                        getattr(profile, "family_address", None) or "",
                        fv,
                    )
                else:
                    value = "-"
            elif field_path == "spouse_wali_district_display":
                if _show_spouse_wali_address_block(profile):
                    fd = getattr(profile, "family_district", None)
                    if fd:
                        value = _safe_name(fd)
                    else:
                        fv = getattr(profile, "family_village", None)
                        reg = (
                            getattr(fv.district, "regency", None)
                            if fv and getattr(fv, "district", None)
                            else None
                        )
                        value = _safe_name(reg)
                else:
                    value = "-"
            elif field_path == "spouse_wali_province_display":
                if _show_spouse_wali_address_block(profile):
                    fp = getattr(profile, "family_province", None)
                    if fp:
                        value = getattr(fp, "name", "-")
                    else:
                        fv = getattr(profile, "family_village", None)
                        reg = (
                            getattr(fv.district, "regency", None)
                            if fv and getattr(fv, "district", None)
                            else None
                        )
                        prov = getattr(reg, "province", None) if reg else None
                        value = getattr(prov, "name", "-") if prov else "-"
                else:
                    value = "-"
            elif field_path == "spouse_wali_postal_code":
                if _show_spouse_wali_address_block(profile):
                    family_village = getattr(profile, "family_village", None)
                    value = _first_non_empty(
                        getattr(profile, "family_postal_code", None),
                        str(getattr(family_village, "code", "") or "").strip(),
                    )
                else:
                    value = "-"
            # Referral & Admin
            elif field_path == "referrer_name":
                referrer = getattr(profile, "referrer", None)
                if referrer:
                    value = getattr(referrer, "full_name", "") or getattr(referrer, "email", "-")
                else:
                    value = "-"
            elif field_path == "verification_status":
                status = _get_nested_value(profile, "verification_status", "")
                value = _format_verification_status(status) if status != "-" else "-"
            elif field_path == "verified_by_name":
                verified_by = getattr(profile, "verified_by", None)
                if verified_by:
                    value = getattr(verified_by, "full_name", "") or getattr(verified_by, "email", "-")
                else:
                    value = "-"
            elif field_path == "verified_at":
                value = _format_date_dmy(getattr(profile, "verified_at", None))
            elif field_path == "verification_notes":
                value = _get_nested_value(profile, "verification_notes")
            elif field_path == "score":
                score = _get_nested_value(profile, "score", "")
                if score and score != "-":
                    try:
                        score_val = float(score)
                        value = str(int(score_val)) if score_val == int(score_val) else str(score_val)
                    except (ValueError, TypeError):
                        value = score
                else:
                    value = "-"
            elif field_path == "religion":
                value = _get_nested_value(profile, "religion")
            elif field_path == "is_active":
                value = "Aktif" if getattr(applicant, "is_active", False) else "Nonaktif"
            elif field_path == "email_verified":
                value = "Ya" if getattr(applicant, "email_verified", False) else "Tidak"
            elif field_path == "created_at":
                created = getattr(profile, "created_at", None) or getattr(applicant, "date_joined", None)
                value = _format_date_dmy(created)
            elif field_path == "passport_expiry_date":
                value = _format_date_dmy(getattr(profile, "passport_expiry_date", None))
            # Work Experiences
            elif field_path == "work_experiences":
                value = _format_work_experiences(profile, request)
            else:
                value = _get_nested_value(profile, field_path)
            
            cell.value = value
        
        # Process document columns - each document type gets its own column
        documents = getattr(profile, "documents", None)
        if documents:
            try:
                docs = list(documents.all()) if hasattr(documents, 'all') else documents
                
                # Create a map of document_type_id -> document
                doc_map = {}
                for doc in docs:
                    doc_type = getattr(doc, "document_type", None)
                    if doc_type:
                        doc_map[doc_type.id] = doc
                
                # Fill in document columns
                for doc_type_id, col_offset in document_type_columns.items():
                    col_idx = col_offset + 1  # columns are 1-indexed
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.alignment = data_alignment
                    
                    if doc_type_id in doc_map:
                        doc = doc_map[doc_type_id]
                        file_field = getattr(doc, "file", None)
                        if file_field:
                            url = _get_file_url(file_field, request)
                            cell.value = url
                        else:
                            cell.value = "-"
                    else:
                        cell.value = "-"
                        
            except Exception as e:
                # If there's an error, just leave document columns empty
                pass
        else:
            # No documents, fill all document columns with "-"
            for doc_type_id, col_offset in document_type_columns.items():
                col_idx = col_offset + 1
                cell = ws.cell(row=row_num, column=col_idx)
                cell.alignment = data_alignment
                cell.value = "-"
        
        row_num += 1
    
    # Auto-adjust column widths
    for col_idx, (label, field_path) in enumerate(all_columns, start=1):
        col_letter = get_column_letter(col_idx)
        # Set minimum width based on header length
        min_width = max(len(label) + 2, 10)
        
        # Special handling for multi-line fields
        if field_path in [
            "work_experiences",
            "verification_notes",
            "address_combined",
            "family_address_combined",
            "spouse_wali_address_combined",
        ]:
            min_width = max(min_width, 40)  # Wider columns for long text
        # Document columns should be wide enough for URLs
        elif field_path.startswith("document_"):
            min_width = max(min_width, 50)  # Wide columns for URLs
        
        ws.column_dimensions[col_letter].width = min_width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
